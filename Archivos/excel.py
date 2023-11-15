#pip install openpyxl

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Hoja1'

sheet['A1'] = 'Hola'
sheet['B1'] = 'Mundo'

wb.save('ejemplo.xlsx')

wb = openpyxl.load_workbook('ejemplo.xlsx')
sheet = wb['Hoja1']

for row in sheet.iter_rows(values_only=True):
    print(row)